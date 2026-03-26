"""
GST Discovery Tool - MCA Updater
Writes discovery results back to MCA master files and marks Gst_Updated.
"""
import logging
import json
import os
from openpyxl import load_workbook
import config

log = logging.getLogger("gst_discovery")

# Column mapping for FICCI & GST.xlsx -> TS GST sheet
# Existing: Imp, Sl.No, Taxpayers Name, GSTIN, Trade Name, Trade Name New File, Pincode, Mobile Number, Address, Name
TSGST_COL_MAP = {
    "mobile_number": 8,    # H (col 8)
    "address": 9,          # I (col 9)
    "upi_name": 10,        # J (col 10)
}


def update_ficci_gst_file(conn, batch_size=100):
    """
    Update FICCI & GST.xlsx TS GST sheet with discovery results.
    Adds mobile number, address, name from discovery.
    Adds Gst_Updated tag to the company name.
    """
    log.info("Updating FICCI & GST file with discovery results...")

    # Load discovery results from db
    results = conn.execute("""
        SELECT gstin, mobile_number, upi_name, address, trade_name, legal_name,
               directors, business_owners, hsn_codes, related_gstins, company_type
        FROM discovery_results
    """).fetchall()

    if not results:
        log.info("No discovery results to update.")
        return 0

    # Build lookup by GSTIN
    result_map = {}
    for r in results:
        result_map[r["gstin"]] = dict(r)

    log.info(f"  {len(result_map)} results to write back")

    # Open FICCI & GST workbook for writing
    wb = load_workbook(config.FICCI_GST_FILE)
    ws = wb[config.FICCI_GST_SHEET]

    # Find header row and column positions
    header_row = None
    gstin_col = None
    mobile_col = None
    address_col = None
    name_col = None
    tag_col = None

    for row in ws.iter_rows(min_row=1, max_row=5):
        for cell in row:
            val = str(cell.value).strip().lower() if cell.value else ""
            if val == "gstin":
                header_row = cell.row
                gstin_col = cell.column
            elif "mobile" in val:
                mobile_col = cell.column
            elif val == "address":
                address_col = cell.column
            elif val == "name" and cell.column > 5:  # avoid "Taxpayers Name"
                name_col = cell.column

    if not gstin_col:
        log.error("Could not find GSTIN column in TS GST sheet")
        return 0

    # Find or create Gst_Updated column
    max_col = ws.max_column
    for col in range(1, max_col + 1):
        val = ws.cell(row=header_row, column=col).value
        if val and "gst_updated" in str(val).lower():
            tag_col = col
            break

    if not tag_col:
        tag_col = max_col + 1
        ws.cell(row=header_row, column=tag_col, value="Gst_Updated")
        log.info(f"  Added Gst_Updated column at position {tag_col}")

    # If mobile/address/name columns don't exist, create them
    if not mobile_col:
        mobile_col = tag_col
        tag_col += 1
        ws.cell(row=header_row, column=mobile_col, value="Mobile Number")
    if not address_col:
        address_col = tag_col
        tag_col += 1
        ws.cell(row=header_row, column=address_col, value="Address")
    if not name_col:
        name_col = tag_col
        tag_col += 1
        ws.cell(row=header_row, column=name_col, value="Name")

    # Update rows
    updated = 0
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
        gstin_cell = row[gstin_col - 1]
        gstin = str(gstin_cell.value).strip() if gstin_cell.value else ""

        if gstin in result_map:
            data = result_map[gstin]

            if data.get("mobile_number"):
                ws.cell(row=gstin_cell.row, column=mobile_col, value=data["mobile_number"])
            if data.get("address"):
                ws.cell(row=gstin_cell.row, column=address_col, value=data["address"])
            if data.get("upi_name"):
                ws.cell(row=gstin_cell.row, column=name_col, value=data["upi_name"])

            # Mark Gst_Updated
            ws.cell(row=gstin_cell.row, column=tag_col, value="Yes")
            updated += 1

    wb.save(config.FICCI_GST_FILE)
    log.info(f"  Updated {updated} rows in FICCI & GST file")
    return updated


def update_mca_master_with_tag(conn):
    """
    Update MCA eir files: add Gst_Updated tag to company names
    that have been discovered via GST.
    """
    log.info("Updating MCA master files with Gst_Updated tags...")

    # Get all discovered companies with CIN
    results = conn.execute("""
        SELECT gstin, cin, trade_name, legal_name, company_type,
               mobile_number, upi_name, directors
        FROM discovery_results
        WHERE cin IS NOT NULL AND cin != ''
    """).fetchall()

    if not results:
        log.info("No CIN-matched results to update in MCA files.")
        return 0

    # Build CIN lookup
    cin_map = {}
    for r in results:
        cin_map[r["cin"]] = dict(r)

    log.info(f"  {len(cin_map)} CIN-matched companies to tag in MCA files")

    total_updated = 0

    # Scan each MCA eir file
    for year_folder in config.MCA_YEAR_FOLDERS:
        folder_path = os.path.join(config.MCA_BASE_DIR, year_folder)
        if not os.path.isdir(folder_path):
            continue

        import glob
        files = glob.glob(os.path.join(folder_path, "*.xlsx"))
        for fpath in sorted(files):
            try:
                wb = load_workbook(fpath)
                ws = wb[wb.sheetnames[0]]  # Indian companies sheet

                # Find CIN column and check for Gst_Updated column
                cin_col = 1  # CIN is always column A
                name_col = 2  # Company Name is column B

                # Find or create Gst_Updated column
                max_col = ws.max_column
                tag_col = None
                header_row = 1

                for col in range(1, max_col + 1):
                    val = ws.cell(row=header_row, column=col).value
                    if val and "gst_updated" in str(val).lower():
                        tag_col = col
                        break

                if not tag_col:
                    tag_col = max_col + 1
                    ws.cell(row=header_row, column=tag_col, value="Gst_Updated")

                file_updates = 0
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    cin = str(row[cin_col - 1].value).strip() if row[cin_col - 1].value else ""
                    if cin in cin_map:
                        ws.cell(row=row[0].row, column=tag_col, value="Yes")
                        file_updates += 1

                if file_updates > 0:
                    wb.save(fpath)
                    total_updated += file_updates
                    log.info(f"  {os.path.basename(fpath)}: tagged {file_updates} companies")
                else:
                    wb.close()

            except Exception as e:
                log.warning(f"  Error updating {os.path.basename(fpath)}: {e}")

    log.info(f"Total MCA companies tagged: {total_updated}")
    return total_updated


def export_discovery_csv(conn, output_path=None):
    """Export all discovery results to CSV for review."""
    import csv

    if not output_path:
        os.makedirs(config.OUTPUT_DIR, exist_ok=True)
        output_path = os.path.join(config.OUTPUT_DIR, "gst_discovery_results.csv")

    results = conn.execute("SELECT * FROM discovery_results").fetchall()
    if not results:
        log.info("No results to export.")
        return

    fields = results[0].keys()
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        for r in results:
            writer.writerow(dict(r))

    log.info(f"Exported {len(results)} results to {output_path}")
