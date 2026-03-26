"""
GST Discovery Tool - Data Loader
Loads GST numbers and MCA company data into SQLite for fast lookup.
"""
import sqlite3
import os
import glob
import logging
from openpyxl import load_workbook
import config

log = logging.getLogger("gst_discovery")


def get_db(db_path=None):
    """Get a connection to the working database."""
    path = db_path or config.WORK_DB
    conn = sqlite3.connect(path)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    return conn


def init_db(conn):
    """Create tables if they don't exist."""
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS gst_numbers (
            gstin TEXT PRIMARY KEY,
            trade_name TEXT,
            division TEXT,
            circle TEXT,
            type TEXT,
            company_type TEXT,       -- 'PVT_LTD', 'LLP', 'REGULAR'
            gst_updated INTEGER DEFAULT 0
        );

        CREATE TABLE IF NOT EXISTS mca_companies (
            cin TEXT PRIMARY KEY,
            company_name TEXT,
            date_of_incorporation TEXT,
            state TEXT,
            roc TEXT,
            category TEXT,
            sub_category TEXT,
            class TEXT,
            authorized_capital TEXT,
            paid_capital TEXT,
            activity_description TEXT,
            registered_address TEXT,
            email TEXT,
            source_file TEXT
        );

        CREATE TABLE IF NOT EXISTS discovery_results (
            gstin TEXT PRIMARY KEY,
            trade_name TEXT,
            legal_name TEXT,
            company_type TEXT,
            cin TEXT,
            mobile_number TEXT,
            mobile_number_tgct TEXT,
            upi_name TEXT,
            address TEXT,
            hsn_codes TEXT,
            directors TEXT,          -- JSON array
            business_owners TEXT,    -- JSON array
            related_gstins TEXT,    -- JSON array from knowyourgst
            source TEXT,
            discovered_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (gstin) REFERENCES gst_numbers(gstin)
        );

        CREATE INDEX IF NOT EXISTS idx_mca_name ON mca_companies(company_name);
        CREATE INDEX IF NOT EXISTS idx_gst_updated ON gst_numbers(gst_updated);
        CREATE INDEX IF NOT EXISTS idx_gst_trade ON gst_numbers(trade_name);
    """)
    conn.commit()


def load_gst_numbers(conn):
    """Load GST numbers from TG GST New.xlsx into SQLite."""
    cursor = conn.cursor()
    existing = cursor.execute("SELECT COUNT(*) FROM gst_numbers").fetchone()[0]
    if existing > 0:
        log.info(f"GST numbers already loaded ({existing} records). Skipping.")
        return existing

    log.info(f"Loading GST numbers from {config.GST_FILE}...")
    wb = load_workbook(config.GST_FILE, read_only=True)

    total = 0
    for sheet_name in [config.GST_SHEET_PRIMARY, config.GST_SHEET_SECONDARY]:
        ws = wb[sheet_name]
        batch = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            # Skip header rows (row 0 = header, row 1 = numbers)
            if i < 2:
                continue
            sno, division, circle, gstin, trade_name, gst_type, remarks = (
                list(row) + [None] * 7
            )[:7]

            if not gstin or not isinstance(gstin, str) or len(gstin) != 15:
                continue

            batch.append((
                gstin.strip(),
                str(trade_name).strip() if trade_name else "",
                str(division).strip() if division else "",
                str(circle).strip() if circle else "",
                str(gst_type).strip() if gst_type else "",
            ))

            if len(batch) >= 5000:
                cursor.executemany(
                    "INSERT OR IGNORE INTO gst_numbers "
                    "(gstin, trade_name, division, circle, type) "
                    "VALUES (?, ?, ?, ?, ?)",
                    batch
                )
                total += len(batch)
                batch = []

        if batch:
            cursor.executemany(
                "INSERT OR IGNORE INTO gst_numbers "
                "(gstin, trade_name, division, circle, type) "
                "VALUES (?, ?, ?, ?, ?)",
                batch
            )
            total += len(batch)

        log.info(f"  Sheet '{sheet_name}': loaded records (running total: {total})")

    conn.commit()
    wb.close()
    log.info(f"Total GST numbers loaded: {total}")
    return total


def load_mca_from_eir(conn):
    """Load MCA company data from monthly eir Excel files (2016-2021)."""
    cursor = conn.cursor()
    existing = cursor.execute("SELECT COUNT(*) FROM mca_companies").fetchone()[0]
    if existing > 0:
        log.info(f"MCA companies already loaded ({existing} records). Skipping.")
        return existing

    log.info("Loading MCA company data from eir files...")
    total = 0

    for year_folder in config.MCA_YEAR_FOLDERS:
        folder_path = os.path.join(config.MCA_BASE_DIR, year_folder)
        if not os.path.isdir(folder_path):
            continue

        files = glob.glob(os.path.join(folder_path, "*.xlsx"))
        for fpath in sorted(files):
            fname = os.path.basename(fpath)
            log.info(f"  Loading {year_folder}/{fname}...")
            try:
                wb = load_workbook(fpath, read_only=True)
                # First sheet = Indian companies registered
                ws = wb[wb.sheetnames[0]]
                batch = []
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i == 0:
                        continue  # skip header
                    vals = list(row) + [None] * 14
                    cin = vals[0]
                    if not cin or not isinstance(cin, str):
                        continue

                    batch.append((
                        cin.strip(),
                        str(vals[1]).strip() if vals[1] else "",
                        str(vals[3]).strip() if vals[3] else "",
                        str(vals[4]).strip() if vals[4] else "",
                        str(vals[5]).strip() if vals[5] else "",
                        str(vals[6]).strip() if vals[6] else "",
                        str(vals[7]).strip() if vals[7] else "",
                        str(vals[8]).strip() if vals[8] else "",
                        str(vals[9]).strip() if vals[9] else "",
                        str(vals[10]).strip() if vals[10] else "",
                        str(vals[12]).strip() if vals[12] else "",
                        str(vals[13]).strip() if vals[13] else "",
                        "",  # email not in eir
                        f"{year_folder}/{fname}",
                    ))

                    if len(batch) >= 5000:
                        cursor.executemany(
                            "INSERT OR IGNORE INTO mca_companies "
                            "(cin, company_name, date_of_incorporation, state, "
                            "roc, category, sub_category, class, "
                            "authorized_capital, paid_capital, "
                            "activity_description, registered_address, "
                            "email, source_file) "
                            "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                            batch
                        )
                        total += len(batch)
                        batch = []

                if batch:
                    cursor.executemany(
                        "INSERT OR IGNORE INTO mca_companies "
                        "(cin, company_name, date_of_incorporation, state, "
                        "roc, category, sub_category, class, "
                        "authorized_capital, paid_capital, "
                        "activity_description, registered_address, "
                        "email, source_file) "
                        "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                        batch
                    )
                    total += len(batch)

                conn.commit()
                wb.close()
            except Exception as e:
                log.warning(f"  Error loading {fname}: {e}")

    # Also load MCA Metros 2015
    if os.path.exists(config.MCA_METROS_FILE):
        log.info("Loading MCA Metros 2015...")
        try:
            wb = load_workbook(config.MCA_METROS_FILE, read_only=True)
            ws = wb[wb.sheetnames[0]]
            batch = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    continue
                vals = list(row) + [None] * 12
                cin = vals[0]
                if not cin or not isinstance(cin, str):
                    continue
                batch.append((
                    cin.strip(),
                    str(vals[1]).strip() if vals[1] else "",
                    str(vals[2]).strip() if vals[2] else "",
                    str(vals[5]).strip() if vals[5] else "",
                    "", "", "", "",
                    str(vals[3]).strip() if vals[3] else "",
                    str(vals[4]).strip() if vals[4] else "",
                    "",
                    str(vals[7]).strip() if vals[7] else "",
                    str(vals[8]).strip() if vals[8] else "",
                    "MCA Metros 2015",
                ))
                if len(batch) >= 5000:
                    cursor.executemany(
                        "INSERT OR IGNORE INTO mca_companies "
                        "(cin, company_name, date_of_incorporation, state, "
                        "roc, category, sub_category, class, "
                        "authorized_capital, paid_capital, "
                        "activity_description, registered_address, "
                        "email, source_file) "
                        "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                        batch
                    )
                    total += len(batch)
                    batch = []
            if batch:
                cursor.executemany(
                    "INSERT OR IGNORE INTO mca_companies "
                    "(cin, company_name, date_of_incorporation, state, "
                    "roc, category, sub_category, class, "
                    "authorized_capital, paid_capital, "
                    "activity_description, registered_address, "
                    "email, source_file) "
                    "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                    batch
                )
                total += len(batch)
            conn.commit()
            wb.close()
        except Exception as e:
            log.warning(f"Error loading MCA Metros: {e}")

    log.info(f"Total MCA companies loaded: {total}")
    return total


def load_all(db_path=None):
    """Load all data sources into SQLite."""
    conn = get_db(db_path)
    init_db(conn)
    gst_count = load_gst_numbers(conn)
    mca_count = load_mca_from_eir(conn)
    log.info(f"Database ready: {gst_count} GSTINs, {mca_count} MCA companies")
    return conn


def get_pending_gstins(conn, limit=None):
    """Get GST numbers that haven't been processed yet."""
    sql = "SELECT gstin, trade_name, company_type FROM gst_numbers WHERE gst_updated = 0"
    if limit:
        sql += f" LIMIT {limit}"
    return conn.execute(sql).fetchall()


def mark_gst_updated(conn, gstin):
    """Mark a GSTIN as processed."""
    conn.execute("UPDATE gst_numbers SET gst_updated = 1 WHERE gstin = ?", (gstin,))
    conn.commit()


def find_mca_company(conn, company_name):
    """Find MCA company by name (fuzzy match)."""
    # Exact match first
    row = conn.execute(
        "SELECT * FROM mca_companies WHERE UPPER(company_name) = UPPER(?)",
        (company_name,)
    ).fetchone()
    if row:
        return dict(row)

    # Partial match
    row = conn.execute(
        "SELECT * FROM mca_companies WHERE UPPER(company_name) LIKE UPPER(?)",
        (f"%{company_name}%",)
    ).fetchone()
    if row:
        return dict(row)

    return None


def save_discovery_result(conn, result):
    """Save discovery result to database."""
    conn.execute("""
        INSERT OR REPLACE INTO discovery_results
        (gstin, trade_name, legal_name, company_type, cin,
         mobile_number, mobile_number_tgct, upi_name, address, hsn_codes,
         directors, business_owners, related_gstins, source)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        result.get("gstin"),
        result.get("trade_name"),
        result.get("legal_name"),
        result.get("company_type"),
        result.get("cin"),
        result.get("mobile_number"),
        result.get("mobile_number_tgct"),
        result.get("upi_name"),
        result.get("address"),
        result.get("hsn_codes"),
        result.get("directors"),
        result.get("business_owners"),
        result.get("related_gstins"),
        result.get("source"),
    ))
    conn.commit()
